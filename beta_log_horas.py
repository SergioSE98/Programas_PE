# -*- coding: utf-8 -*-
"""
Created on Wed Apr 14 13:11:26 2021

@author: Sergi
"""

#En primer lugar se importan los módulos a utilizar. 

from openpyxl import Workbook
import numpy as np
import pandas as pd
import datetime
import matplotlib.pyplot as plt
import scipy.stats as scpstats

#A continuación se leen los csv haciendo uso de pandas.

gravim_data = pd.read_csv("filtros_hvs_20_4_10-intervalo_20200601-20201231_SCO_.csv", header=0, sep = ", ", engine = "python", )

beta_data = pd.read_csv("202006_202012_Beta_Raw_Data.csv", header=0)

meteo_data = pd.read_csv("202006_202012_Meteo_SCO.csv", header=0)


#Empiezo realizando ajustes a las fechas. Necesito pasar las fechas como vienen a formato datetime para operar con ellas.

inicio_gravim = gravim_data["Inicio muestreo"]  #Asigno dias de inicio de medida de gravimetrico a una variable
inicio_gravim = pd.to_datetime(inicio_gravim)   #Convierto esa variable a datetime
gravim_data["Inicio muestreo"] = inicio_gravim  #Vuelvo a meter los días de inicio en el dataframe ahora como datetimes


#Este programa es exactamente análogo al anterior, la diferencia radica en que voy a quitar los periodos en que se produjeron
#problemas, que vienen indicados en el logbook. Para ello, llamo a mi variable timestampx (que es como viene en logbook)
#Y creo una mask con lo que me marca el logbook.

timestampx = beta_data["timestampx"]
mask = (timestampx>'2020-06-01 00:00:00')&(timestampx<'2020-07-15 00:00:00')|(timestampx>'2020-07-15 18:00:00')&(timestampx<'2020-07-20 09:00:00')|(timestampx>'2020-07-20 18:00:00')&(timestampx<'2020-08-31 00:00:00')|(timestampx>'2020-09-01 17:30:00')&(timestampx<'2020-11-27 14:00:00')|(timestampx>'2020-11-28 14:00:00')&(timestampx<'2021-01-01 00:00:00')
beta_data = beta_data[mask]  #Aplico la mask, que restringe los días del logbook

#Y ahora ya puedo continuar con el programa como antes, restando las 8 horas etc.

fechas_beta = beta_data["timestampx"]           #Asigno fechas de medidas de beta a una variable
fechas_beta = pd.to_datetime(fechas_beta)       #Convierto a datetime
horas = datetime.timedelta(hours=8)             #Creo un delta de 8 horas que restaré a mis medidas de beta
fechas_beta = fechas_beta - horas               #Resto 8 horas a mis fechas de medidas de beta
beta_data["timestampx"] = fechas_beta           #Meto los nuevos valores de fechas en el dataframe (ahora como datetimes)

#Haciendo esto, consigo que las medidas de atenuación beta hechas a las 8:00 figuren hechas a las 0:00, de forma que pueda
#considerar los días en que se realizan estas únicamente llamando a la fecha sin tener en cuenta las horas.
#Ahora en vez de medidas de 8:00 a 8:00 divido en medidas de 0:00 a 0:00, aunque son las mismas que de 8:00 a 8:00,
#de forma que pueda relacionar la fecha de inicio de gravim con la fecha en que se han tomado las medidas de atenuacion beta
#ese mismo día de inicio, se supone que de 0:00 a 0:00 del día siguiente, aunque en realidad es de 8:00 a 8:00, como
#las de gravimetrico. 

#Ahora genero un nuevo dataframe que contenga las columnas de los dataframes de gravimetrico y atenuación beta, y que
#me escoja solo los días en que hay medidas de gravimetrico y beta attenuation. Al haber hecho la traslacion horaria,
#los valores que hagan merge serán los tomados tanto por gravimetrico como beta desde "Inicio muestreo" a las 8:00 hasta
#"Fin muestreo" a las 8:00

merge = beta_data.merge(gravim_data,left_on = beta_data["timestampx"].dt.date, right_on = gravim_data["Inicio muestreo"].dt.date)

#Ya tengo mi nuevo dataframe, ahora saco de él algunas variables que me interesará usar:

hours_beta = merge["timestampx"].dt.hour     #Añadiendo ".dt.hour" considero solo las horas, por eso era necesario usar datetimes.
dates_beta = merge["timestampx"].dt.date     #Añadiendo "dt.date" considero solo las fechas.

pm_beta = merge["pm"]
pm_gravim = merge["PM"]


#Ahora voy a calcular el valor medio de las medidas de PM10 realizadas por atenuación beta para cada día en que fueron tomadas.
#Adicionalmente calcularé las desviaciones estándar de las medidas que se realizaron esos días.

pm_gravim_no_repetidos = ([])   #Creo arrays vacíos para introducir los calculos
medias_beta = ([])
desviaciones_beta = ([])             #Esto es la media a lo largo de un dia de las desviaciones de medidas horarias
desviaciones_beta_dia = ([])         #Esto es la desviacion estandar de las medidas en un dia
y_hour=([])                          #Aquí acumulo el pm en cada hora
media_hora=([])                      #Media de los pm cada hora
desv_hora=([])                       #Desv de los pm cada hora


#Creo un bucle donde evaluo si mi pm en el gravimetrico cambia, cuando cambia es porque ha cambiado el día
#entonces añado la media de valores de pm de beta medidos ese día, las desviaciones estandar, y el valor de gravimetrico
#Así consigo medidas de PM10 en Gravim y beta para cada día. 

#Es necesario añadir una condición para tratar el último dato, si no se llegaría a un error. Basicamente se hace todo lo
#que se ha hecho para los anteriores datos y se acaba el bucle.

#En este caso, creo previa a la condición de cambio de fecha, otra condición para ver si se ha cambiado de hora,
#y así poder acumular los datos en horas separadas y poder obtener medias horarias.

rango = (len(merge))  #Rango de actuación del bucle

for j in range(rango):
        
    if j == (rango-1):                              #Esto lo añado para evitar un error en el último término.
        y_hour.append(pm_beta[j])                           #Basicamente hago todo lo que se ha ido haciendo.
        media_hora.append(np.mean(y_hour))
        desv_hora.append(np.std(y_hour))
        medias_beta.append(np.mean(media_hora))             
        desviaciones_beta.append(np.std(media_hora))
        desviaciones_beta_dia.append(np.std(media_hora))
        pm_gravim_no_repetidos.append(pm_gravim[j])
        break    
    if hours_beta[j] == hours_beta[j+1]:                    #Mientras la hora sea la misma, acumulo los valores de pm  
        y_hour.append(pm_beta[j])                           #Los voy añadiendo a ese array.
        #continue
    
    if hours_beta[j] != hours_beta[j+1]:                    #Así consigo diferenciar cuando se produce un cambio de hora
        media_hora.append(np.mean(y_hour))                  #Entonces calculo media de valores en una hora
        desv_hora.append(np.std(y_hour))                    #Y también desviación estándar de esos valores en una hora
        y_hour=([])                                         #Por último reseteo este array para acumular hora siguiente.
        
    if dates_beta[j] != dates_beta[j+1]:                    #Así consigo diferenciar cuando se produce un cambio de día.
        medias_beta.append(np.mean(media_hora))             #Procedo como antes, pero ahora para la media en todo un día.
        desviaciones_beta.append(np.mean(desv_hora))        #Esto da la media de las desv de cada hora 
        desviaciones_beta_dia.append(np.std(media_hora))    #Esto da la desv de las medias de cada hora a lo largo del dia
        pm_gravim_no_repetidos.append(pm_gravim[j])         #Con esto acumulo también el dato de gravim de ese día
        media_hora = ([])                                   #Y reseteo los array.
        desv_hora = ([])
        #continue


#print(pm_gravim_no_repetidos)        #Puedo desmarcando esos corchetes imprimir los valores de gravim y beta para cada día.
#print(medias_beta)

#Ahora ya puedo hacer un ajuste lineal de mis datos, comparando los de gravimetrico con las medias de PM medido por
#atenuación beta en períodos iguales (que he acotado haciendo el "merge").

ajuste = scpstats.linregress(pm_gravim_no_repetidos,medias_beta)   #Esto realiza el ajuste lineal de mis datos

#print(ajuste)       #Puedo imprimir los valores obtenidos, aunque los añadiré a mi gráfica.

#Calculo las medidas de "y"  usando mi ajuste y los datos de gravimétrico (mi x). Esto lo uso para representar el ajuste graficamente.

y_ajuste=([])
for i in  range(len(pm_gravim_no_repetidos)):                  
    calculo = ajuste[0]*pm_gravim_no_repetidos[i] + ajuste[1]
    y_ajuste.append(calculo)

#Ya puedo crear mi figura y representar lo obtenido, considerando también barras de error (usando la desv estándar)
    
#Recalcar que se pueden representar las desviaciones estándar de las medias horarias a lo largo de un día "desviaciones_beta_dia",
#O las medias de las desviaciones estandar de cada hora (para cada día) "desviaciones_beta".
    
plt.figure()  
plt.errorbar(pm_gravim_no_repetidos, medias_beta, desviaciones_beta_dia, linestyle = "None", marker = ".", label = "Datos (Errorbars = Desviación estándar BAM)")
plt.plot(pm_gravim_no_repetidos, y_ajuste, label= "Ajuste ($a=%.3f, b=%.3f,r^2= %.3f $ )" %(ajuste[0],ajuste[1],ajuste[2]))
plt.xlabel("PM10 medido por el método de referencia (gravimétrico) [$\mu g / m^3$]")
plt.ylabel(r"PM10 medido por el método automático (BAM) [$\mu g / m^3$]")
plt.xlim(0,100)
plt.ylim(0,100)
plt.legend()
plt.grid(linewidth=0.3)
plt.title(r"PM10 medido por BAM vs PM10 medido por gravimétrico ")
plt.savefig("Verificación_corrección_BAM_definitivo.png")


#Si quiero sacar en un excel los valores de PM10 por beta y gravimétrico para los distintos días, usar lo siguiente.
    
wb = Workbook()
ruta = 'salida.xlsx'

hoja = wb.active
hoja.title = "PM_beta"

fila = 1 #Fila donde empezamos
col_pm_beta = 1 #Columna donde guardo los valores
col_pm_grav = 2

for pm_beta, pm_grav in zip(medias_beta,pm_gravim_no_repetidos):  #Aquí pongo en "in" el valor que quiero sacar como columna
    hoja.cell(column=col_pm_beta, row=fila, value=pm_beta)
    hoja.cell(column=col_pm_grav, row=fila, value=pm_grav)
    fila+=1

wb.save(filename = ruta)
